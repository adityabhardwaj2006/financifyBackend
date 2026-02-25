"""
Excel Balance Sheet Parser.

Handles diverse Excel balance sheet formats including:
- Schedule III (Companies Act, 2013) format
- Traditional T-account format (Dr/Cr sides)
- Generic two-column (Label, Value) layouts
- Multi-sheet workbooks
- MULTI-YEAR columns with automatic detection and extraction

The parser auto-detects which columns contain labels and which contain
numeric values, handles section headers, subtotals, and merged regions.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from financial_mapper.logging_setup import get_logger

logger = get_logger("excel_parser")

# Rows whose normalised text matches these patterns are section headers,
# not actual data rows.  We skip them during extraction.
_HEADER_PATTERNS: list[re.Pattern] = [
    re.compile(r"^(i+\.|iv\.|v\.)\s", re.IGNORECASE),  # Roman numeral headings
    re.compile(r"^\d+\)\s"),  # Numbered headings like "1)"
    re.compile(r"^(statement of|balance sheet|profit and loss|trading account|"
               r"for the year|as per schedule|particulars|dr\.|cr\.|"
               r"equity and liabilities|assets|expenses|ratios)", re.IGNORECASE),
]

# Labels that are clearly totals/subtotals — we keep them as they may map
# to canonical fields like "Total Assets", "Total Liabilities" etc.
_SKIP_PATTERNS: list[re.Pattern] = [
    re.compile(r"^(note no|notes?)$", re.IGNORECASE),
]


def _is_section_header(text: str) -> bool:
    """Return True if the text looks like a section header, not a data row."""
    stripped = text.strip()
    if not stripped:
        return True
    for pat in _HEADER_PATTERNS:
        if pat.search(stripped):
            return True
    return False


def _should_skip(text: str) -> bool:
    """Return True if the label should be entirely skipped."""
    stripped = text.strip()
    for pat in _SKIP_PATTERNS:
        if pat.match(stripped):
            return True
    return False


def _clean_label(text: str) -> str:
    """Clean a raw label extracted from Excel."""
    if not text:
        return ""
    s = str(text).strip()
    # Remove leading roman-numeral-style prefixes like "I. ", "IV. ", "IX. "
    s = re.sub(r"^[IVXLC]+\.\s*", "", s, flags=re.IGNORECASE)
    # Remove leading numbering like "1. ", "2. ", "a) ", "b) "
    s = re.sub(r"^\d+[\.\)]\s*", "", s)
    s = re.sub(r"^[a-f]\)\s*", "", s, flags=re.IGNORECASE)
    # Remove leading indicators like "To ", "By "
    s = re.sub(r"^(to|by)\s+", "", s, flags=re.IGNORECASE)
    # Remove leading whitespace/indentation artifacts
    s = s.strip()
    # Remove trailing colons
    s = s.rstrip(":")
    return s.strip()


def _is_numeric(val: Any) -> bool:
    """Check if a value is numeric (int, float) and not a date."""
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, datetime):
        return False
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "").replace("₹", "").replace("$", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = cleaned[1:-1]
        try:
            float(cleaned)
            return True
        except (ValueError, TypeError):
            return False
    return False


def _to_number(val: Any) -> Optional[float]:
    """Convert a value to a float, returning None if not possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, datetime):
        return None
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "").replace("₹", "").replace("$", "")
        neg = False
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = cleaned[1:-1]
            neg = True
        try:
            result = float(cleaned)
            return -result if neg else result
        except (ValueError, TypeError):
            return None
    return None


def _is_label(val: Any) -> bool:
    """Check if a value looks like a text label (not a number/date)."""
    if val is None:
        return False
    if isinstance(val, datetime):
        return False
    if isinstance(val, (int, float)):
        return False
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return False
        # If it parses as a number, it's not a label
        cleaned = s.replace(",", "").replace("₹", "").replace("$", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = cleaned[1:-1]
        try:
            float(cleaned)
            return False
        except (ValueError, TypeError):
            return True
    return False


def _extract_year_from_header(cell: Any) -> Optional[str]:
    """Extract year identifier from a header cell (date, year number, or text)."""
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    
    if isinstance(cell, (int, float)):
        year_num = int(cell)
        if 2000 <= year_num <= 2100:
            return str(year_num)
    
    if isinstance(cell, str):
        # Try to extract year from various formats:
        # "31-03-2025", "2025-03-31", "FY 2025", "FY2025", "Year 2025", etc.
        text = cell.strip()
        
        # Pattern 1: DD-MM-YYYY or DD/MM/YYYY
        match = re.search(r'\b(\d{2})[-/](\d{2})[-/](\d{4})\b', text)
        if match:
            return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        
        # Pattern 2: YYYY-MM-DD or YYYY/MM/DD
        match = re.search(r'\b(\d{4})[-/](\d{2})[-/](\d{2})\b', text)
        if match:
            return match.group(0).replace('/', '-')
        
        # Pattern 3: Just a 4-digit year
        match = re.search(r'\b(20\d{2})\b', text)
        if match:
            return match.group(1)
        
        # Pattern 4: Fiscal year notation
        match = re.search(r'FY\s*(\d{4})', text, re.IGNORECASE)
        if match:
            return f"FY{match.group(1)}"
    
    return None


class ExcelParser:
    """Parse Excel workbooks containing financial statements.

    Supports auto-detection of:
    - Which columns contain labels vs values
    - Multiple sections within a single sheet (P&L, Balance Sheet, Trading A/c)
    - T-account (Dr/Cr) layouts
    - Schedule III format with multi-year columns
    - AUTOMATIC extraction of ALL year columns
    """

    def __init__(self, year_index: Optional[int] = None) -> None:
        """
        Parameters
        ----------
        year_index:
            When multiple year-columns exist, which one to pick.
            0 = first (leftmost) year column, 1 = second, etc.
            If None, will extract ALL years.
        """
        self.year_index = year_index

    def parse_file(self, path: Union[str, Path]) -> Union[List[Tuple[str, Any]], Dict[str, List[Tuple[str, Any]]]]:
        """Parse an Excel file and return (label, value) pairs.

        Returns
        -------
        If year_index is specified: List[Tuple[str, Any]]
            List of (label, value) pairs for that specific year
        
        If year_index is None: Dict[str, List[Tuple[str, Any]]]
            Dictionary mapping year identifiers to lists of (label, value) pairs
            Example: {"2025": [(label1, val1), ...], "2026": [(label2, val2), ...]}
        """
        path = Path(path)
        wb = openpyxl.load_workbook(path, data_only=True)
        
        if self.year_index is not None:
            # Legacy mode: extract single year
            all_pairs: List[Tuple[str, Any]] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info("Parsing sheet: %s (%d rows × %d cols)",
                            sheet_name, ws.max_row, ws.max_column)
                pairs = self._parse_sheet(ws, sheet_name)
                all_pairs.extend(pairs)
                logger.info("Extracted %d pairs from sheet '%s'", len(pairs), sheet_name)

            wb.close()
            
            # Deduplicate
            seen: Dict[str, Any] = {}
            deduped: List[Tuple[str, Any]] = []
            for label, value in all_pairs:
                key = label.strip().lower()
                if key in seen:
                    logger.debug("Duplicate label across sheets: %r", label)
                seen[key] = value
                deduped.append((label, value))

            logger.info("Total extracted pairs: %d (from %d sheets)",
                         len(deduped), len(wb.sheetnames))
            return deduped
        
        else:
            # NEW MODE: Extract ALL years
            year_data: Dict[str, List[Tuple[str, Any]]] = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info("Parsing sheet for ALL years: %s (%d rows × %d cols)",
                            sheet_name, ws.max_row, ws.max_column)
                sheet_year_data = self._parse_sheet_multi_year(ws, sheet_name)
                
                # Merge data from this sheet into year_data
                for year, pairs in sheet_year_data.items():
                    if year not in year_data:
                        year_data[year] = []
                    year_data[year].extend(pairs)
                    logger.info("Extracted %d pairs for year '%s' from sheet '%s'", 
                              len(pairs), year, sheet_name)
            
            wb.close()
            
            # Deduplicate within each year
            for year in year_data:
                seen: Dict[str, Any] = {}
                deduped: List[Tuple[str, Any]] = []
                for label, value in year_data[year]:
                    key = label.strip().lower()
                    if key not in seen:
                        deduped.append((label, value))
                        seen[key] = value
                year_data[year] = deduped
                logger.info("Year '%s': %d unique pairs after deduplication", year, len(deduped))
            
            return year_data

    def _parse_sheet_multi_year(self, ws: Worksheet, sheet_name: str) -> Dict[str, List[Tuple[str, Any]]]:
        """Parse a single worksheet and extract data for ALL years."""
        # Read all cells into a grid
        grid: List[List[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column, values_only=True):
            grid.append(list(row))

        if not grid:
            return {}

        # Detect the layout strategy
        layout = self._detect_layout(grid)
        logger.info("Sheet '%s' detected layout: %s", sheet_name, layout["type"])

        if layout["type"] == "schedule_iii" and "value_cols" in layout:
            # Extract year headers
            year_headers = self._extract_year_headers(grid, layout)
            return self._parse_schedule_iii_multi_year(grid, layout, year_headers)
        elif layout["type"] == "generic" and "value_cols" in layout:
            year_headers = self._extract_year_headers(grid, layout)
            return self._parse_generic_multi_year(grid, layout, year_headers)
        else:
            # Fallback: use legacy single-year parsing for first column
            pairs = self._parse_sheet_legacy(ws, sheet_name, grid, layout)
            return {"Unknown Year": pairs}

    def _extract_year_headers(self, grid: List[List[Any]], layout: Dict[str, Any]) -> Dict[int, str]:
        """Extract year identifiers from column headers.
        
        Returns
        -------
        Dict mapping column index to year identifier
        """
        value_cols = layout.get("value_cols", [])
        header_row_idx = layout.get("header_row", 0)
        
        year_headers: Dict[int, str] = {}
        
        # Check the header row and a few rows above/below it
        for row_offset in [0, -1, 1, -2]:
            row_idx = header_row_idx + row_offset
            if 0 <= row_idx < len(grid):
                row = grid[row_idx]
                for col_idx in value_cols:
                    if col_idx < len(row) and col_idx not in year_headers:
                        cell = row[col_idx]
                        year = _extract_year_from_header(cell)
                        if year:
                            year_headers[col_idx] = year
                            logger.info("Detected year '%s' in column %d (row %d)", 
                                      year, col_idx, row_idx)
        
        # If no years detected, create default labels
        if not year_headers:
            for i, col_idx in enumerate(value_cols):
                year_headers[col_idx] = f"Year {i + 1}"
                logger.warning("No year header found for column %d, using '%s'", 
                             col_idx, year_headers[col_idx])
        
        return year_headers

    def _parse_schedule_iii_multi_year(
        self, grid: List[List[Any]], layout: Dict[str, Any], year_headers: Dict[int, str]
    ) -> Dict[str, List[Tuple[str, Any]]]:
        """Parse Schedule III format and extract data for all years."""
        label_col = layout["label_col"]
        header_row = layout.get("header_row", 0)
        
        # Initialize result dictionary
        year_data: Dict[str, List[Tuple[str, Any]]] = {year: [] for year in year_headers.values()}
        
        for i, row in enumerate(grid):
            if i <= header_row:
                continue

            # Get the label
            raw_label = row[label_col] if label_col < len(row) else None
            if not _is_label(raw_label):
                continue

            label_text = str(raw_label).strip()
            if _should_skip(label_text):
                continue

            cleaned = _clean_label(label_text)
            if not cleaned or _is_section_header(cleaned):
                continue

            # Extract values for each year column
            for col_idx, year in year_headers.items():
                value = row[col_idx] if col_idx < len(row) else None
                
                # Try adjacent columns if value not found
                if not _is_numeric(value):
                    for offset in [1, -1]:
                        alt_col = col_idx + offset
                        if 0 <= alt_col < len(row) and alt_col != label_col:
                            if _is_numeric(row[alt_col]):
                                value = row[alt_col]
                                break

                if _is_numeric(value):
                    num = _to_number(value)
                    if num is not None:
                        year_data[year].append((cleaned, num))

        return year_data

    def _parse_generic_multi_year(
        self, grid: List[List[Any]], layout: Dict[str, Any], year_headers: Dict[int, str]
    ) -> Dict[str, List[Tuple[str, Any]]]:
        """Parse generic format and extract data for all years."""
        label_col = layout.get("label_col", 0)
        
        # Initialize result dictionary
        year_data: Dict[str, List[Tuple[str, Any]]] = {year: [] for year in year_headers.values()}
        
        for row in grid:
            raw_label = row[label_col] if label_col < len(row) else None
            
            if not _is_label(raw_label):
                continue
            
            label_text = str(raw_label).strip()
            if _should_skip(label_text):
                continue
            
            cleaned = _clean_label(label_text)
            if not cleaned or _is_section_header(cleaned):
                continue
            
            # Extract values for each year column
            for col_idx, year in year_headers.items():
                value = row[col_idx] if col_idx < len(row) else None
                
                if _is_numeric(value):
                    num = _to_number(value)
                    if num is not None:
                        year_data[year].append((cleaned, num))
        
        return year_data

    def _parse_sheet_legacy(self, ws: Worksheet, sheet_name: str, grid: List[List[Any]], layout: Dict[str, Any]) -> List[Tuple[str, Any]]:
        """Legacy single-year parsing (backwards compatibility)."""
        if layout["type"] == "t_account":
            return self._parse_t_account(grid, layout)
        elif layout["type"] == "schedule_iii":
            return self._parse_schedule_iii(grid, layout)
        else:
            return self._parse_generic(grid, layout)

    def _parse_sheet(self, ws: Worksheet, sheet_name: str) -> List[Tuple[str, Any]]:
        """Parse a single worksheet (legacy single-year mode)."""
        # Read all cells into a grid
        grid: List[List[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column, values_only=True):
            grid.append(list(row))

        if not grid:
            return []

        # Detect the layout strategy
        layout = self._detect_layout(grid)
        logger.info("Sheet '%s' detected layout: %s", sheet_name, layout["type"])

        if layout["type"] == "t_account":
            return self._parse_t_account(grid, layout)
        elif layout["type"] == "schedule_iii":
            return self._parse_schedule_iii(grid, layout)
        else:
            return self._parse_generic(grid, layout)

    def _detect_layout(self, grid: List[List[Any]]) -> Dict[str, Any]:
        """Detect the sheet layout by analysing column content patterns."""
        max_cols = max(len(row) for row in grid) if grid else 0

        # Check for T-account indicators (Dr/Cr pattern)
        for row in grid[:10]:
            row_text = " ".join(str(c) for c in row if c is not None).lower()
            if ("dr." in row_text and "cr." in row_text) or \
               ("dr. particulars" in row_text) or \
               ("particulars (dr.)" in row_text):
                # T-account: find the debit and credit column groups
                return self._detect_t_account_columns(grid)

        # Check for Schedule III pattern (Particulars + multi-year columns)
        for i, row in enumerate(grid[:5]):
            date_cols = []
            for j, cell in enumerate(row):
                if isinstance(cell, datetime):
                    date_cols.append(j)
                elif isinstance(cell, (int, float)) and 2000 <= cell <= 2100:
                    date_cols.append(j)
            if len(date_cols) >= 1:
                # Find the label column (first column with text)
                label_col = 0
                for j, cell in enumerate(row):
                    if _is_label(cell) and str(cell).strip().lower() in ("particulars",):
                        label_col = j
                        break
                return {
                    "type": "schedule_iii",
                    "label_col": label_col,
                    "value_cols": date_cols,
                    "header_row": i,
                }

        # Generic: find columns with most labels vs most numbers
        return self._detect_generic_columns(grid)

    def _detect_t_account_columns(self, grid: List[List[Any]]) -> Dict[str, Any]:
        """Detect column layout for T-account format."""
        # Find the header row with Dr/Cr indicators
        header_row = 0
        for i, row in enumerate(grid[:10]):
            row_text = " ".join(str(c) for c in row if c is not None).lower()
            if "particulars" in row_text and "amount" in row_text:
                header_row = i
                break

        # Scan columns to find label/value pairs on both sides
        # Typically: Dr label col, Dr value col, Cr label col, Cr value col
        max_cols = max(len(row) for row in grid) if grid else 0
        col_label_count = [0] * max_cols
        col_num_count = [0] * max_cols

        for row in grid[header_row + 1:]:
            for j, cell in enumerate(row):
                if j < max_cols:
                    if _is_label(cell):
                        col_label_count[j] += 1
                    elif _is_numeric(cell):
                        col_num_count[j] += 1

        # Find pairs of (label_col, value_col) — label col has most labels,
        # adjacent col has most numbers
        groups: List[Dict[str, int]] = []
        j = 0
        while j < max_cols - 1:
            if col_label_count[j] > 2:
                # Find the nearest numeric column
                for k in range(j + 1, min(j + 3, max_cols)):
                    if col_num_count[k] > 2:
                        groups.append({"label_col": j, "value_col": k})
                        j = k + 1
                        break
                else:
                    j += 1
            else:
                j += 1

        return {
            "type": "t_account",
            "groups": groups,
            "header_row": header_row,
        }

    def _detect_generic_columns(self, grid: List[List[Any]]) -> Dict[str, Any]:
        """Detect generic two-column layout."""
        max_cols = max(len(row) for row in grid) if grid else 0
        col_label_count = [0] * max_cols
        col_num_count = [0] * max_cols

        for row in grid:
            for j, cell in enumerate(row):
                if j < max_cols:
                    if _is_label(cell):
                        col_label_count[j] += 1
                    elif _is_numeric(cell):
                        col_num_count[j] += 1

        # Find the column with the most labels → label column
        label_col = 0
        if col_label_count:
            label_col = col_label_count.index(max(col_label_count))

        # Find numeric columns (to the right of or adjacent to label col)
        value_cols = []
        for j in range(max_cols):
            if col_num_count[j] > 2 and j != label_col:
                value_cols.append(j)

        if not value_cols:
            # Fallback: use column right of label
            value_cols = [label_col + 1] if label_col + 1 < max_cols else [label_col]

        return {
            "type": "generic",
            "label_col": label_col,
            "value_cols": value_cols,
        }

    def _parse_schedule_iii(
        self, grid: List[List[Any]], layout: Dict[str, Any]
    ) -> List[Tuple[str, Any]]:
        """Parse Schedule III format sheet."""
        label_col = layout["label_col"]
        value_cols = layout["value_cols"]
        header_row = layout.get("header_row", 0)

        # Pick the year column based on year_index
        if self.year_index < len(value_cols):
            val_col = value_cols[self.year_index]
        else:
            val_col = value_cols[0] if value_cols else label_col + 1

        pairs: List[Tuple[str, Any]] = []

        # The sheet may have multiple sections (P&L then Balance Sheet)
        # Each section may re-define column positions
        # Track when value columns shift (Balance Sheet section often uses
        # different columns than P&L section)
        current_val_col = val_col

        for i, row in enumerate(grid):
            if i <= header_row:
                continue

            # Get the label from label_col
            raw_label = row[label_col] if label_col < len(row) else None
            if not _is_label(raw_label):
                # Check if this row signals a section change (e.g. "BALANCE SHEET")
                # which might change the value column layout
                if raw_label and isinstance(raw_label, str):
                    text = raw_label.strip().upper()
                    if "BALANCE SHEET" in text or "EQUITY AND LIABILITIES" in text or "ASSETS" in text:
                        # Look ahead for a new header row with dates
                        for k in range(i, min(i + 5, len(grid))):
                            new_val_cols = []
                            for j, cell in enumerate(grid[k]):
                                if isinstance(cell, datetime) or \
                                   (isinstance(cell, (int, float)) and 2000 <= cell <= 2100):
                                    new_val_cols.append(j)
                            if new_val_cols:
                                idx = min(self.year_index, len(new_val_cols) - 1)
                                current_val_col = new_val_cols[idx]
                                logger.debug("Section change at row %d, new val_col=%d", i, current_val_col)
                                break
                continue

            label_text = str(raw_label).strip()
            if _should_skip(label_text):
                continue

            # Get the value from the current value column
            value = row[current_val_col] if current_val_col < len(row) else None

            # If no value at expected column, scan adjacent columns
            if not _is_numeric(value):
                for offset in [1, -1, 2]:
                    alt_col = current_val_col + offset
                    if 0 <= alt_col < len(row) and alt_col != label_col:
                        if _is_numeric(row[alt_col]):
                            value = row[alt_col]
                            break

            if _is_numeric(value):
                cleaned = _clean_label(label_text)
                if cleaned and not _is_section_header(cleaned):
                    num = _to_number(value)
                    if num is not None:
                        pairs.append((cleaned, num))

        return pairs

    def _parse_t_account(
        self, grid: List[List[Any]], layout: Dict[str, Any]
    ) -> List[Tuple[str, Any]]:
        """Parse T-account (Dr/Cr) format."""
        groups = layout.get("groups", [])
        header_row = layout.get("header_row", 0)
        pairs: List[Tuple[str, Any]] = []

        if not groups:
            logger.warning("T-account layout detected but no column groups found")
            return self._parse_generic(grid, self._detect_generic_columns(grid))

        for row_idx, row in enumerate(grid):
            if row_idx <= header_row:
                continue

            for group in groups:
                lc = group["label_col"]
                vc = group["value_col"]

                raw_label = row[lc] if lc < len(row) else None
                raw_value = row[vc] if vc < len(row) else None

                if _is_label(raw_label) and _is_numeric(raw_value):
                    cleaned = _clean_label(str(raw_label).strip())
                    if cleaned and not _is_section_header(cleaned) and not _should_skip(cleaned):
                        num = _to_number(raw_value)
                        if num is not None and num != 0:
                            pairs.append((cleaned, num))

        # Also scan for the Balance Sheet section which uses a different layout
        # (Liabilities on left, Assets on right)
        bs_start = None
        for i, row in enumerate(grid):
            row_text = " ".join(str(c) for c in row if c is not None).upper()
            if "BALANCE SHEET" in row_text:
                bs_start = i
                break

        if bs_start is not None:
            bs_pairs = self._parse_balance_sheet_section(grid, bs_start)
            pairs.extend(bs_pairs)

        return pairs

    def _parse_balance_sheet_section(
        self, grid: List[List[Any]], start_row: int
    ) -> List[Tuple[str, Any]]:
        """Parse a Balance Sheet section within a T-account sheet."""
        pairs: List[Tuple[str, Any]] = []

        # Re-detect columns for the BS section
        bs_grid = grid[start_row:]
        layout = self._detect_generic_columns(bs_grid)

        # Find column groups in the BS section
        max_cols = max(len(r) for r in bs_grid) if bs_grid else 0
        col_label_count = [0] * max_cols
        col_num_count = [0] * max_cols

        for row in bs_grid:
            for j, cell in enumerate(row):
                if j < max_cols:
                    if _is_label(cell):
                        col_label_count[j] += 1
                    elif _is_numeric(cell):
                        col_num_count[j] += 1

        # Find all (label_col, value_col) groups
        groups: List[Dict[str, int]] = []
        j = 0
        while j < max_cols - 1:
            if col_label_count[j] >= 2:
                for k in range(j + 1, min(j + 3, max_cols)):
                    if col_num_count[k] >= 2:
                        groups.append({"label_col": j, "value_col": k})
                        j = k + 1
                        break
                else:
                    j += 1
            else:
                j += 1

        # Skip header rows (1-2 rows after "BALANCE SHEET")
        data_start = 2
        for i, row in enumerate(bs_grid[1:5], 1):
            row_text = " ".join(str(c) for c in row if c is not None).lower()
            if "amount" in row_text or "liabilities" in row_text:
                data_start = i + 1
                break

        for row in bs_grid[data_start:]:
            for group in groups:
                lc = group["label_col"]
                vc = group["value_col"]

                raw_label = row[lc] if lc < len(row) else None
                raw_value = row[vc] if vc < len(row) else None

                if _is_label(raw_label) and _is_numeric(raw_value):
                    cleaned = _clean_label(str(raw_label).strip())
                    if cleaned and not _is_section_header(cleaned) and not _should_skip(cleaned):
                        # Skip "Total" rows that are just sums, but keep
                        # meaningful totals
                        text_lower = cleaned.lower()
                        if text_lower == "total":
                            continue
                        num = _to_number(raw_value)
                        if num is not None and num != 0:
                            pairs.append((cleaned, num))

        return pairs

    def _parse_generic(
        self, grid: List[List[Any]], layout: Dict[str, Any]
    ) -> List[Tuple[str, Any]]:
        """Parse a generic two-column layout."""
        label_col = layout.get("label_col", 0)
        value_cols = layout.get("value_cols", [1])

        # Pick the value column based on year_index
        if self.year_index < len(value_cols):
            val_col = value_cols[self.year_index]
        else:
            val_col = value_cols[0] if value_cols else 1

        pairs: List[Tuple[str, Any]] = []

        for row in grid:
            raw_label = row[label_col] if label_col < len(row) else None
            raw_value = row[val_col] if val_col < len(row) else None

            if _is_label(raw_label) and _is_numeric(raw_value):
                cleaned = _clean_label(str(raw_label).strip())
                if cleaned and not _is_section_header(cleaned) and not _should_skip(cleaned):
                    num = _to_number(raw_value)
                    if num is not None:
                        pairs.append((cleaned, num))

        return pairs
